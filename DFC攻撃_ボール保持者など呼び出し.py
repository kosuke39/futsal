# -*- coding: utf-8 -*-
"""
Created on Fri Nov 29 13:46:55 2019

@author: 戸田　康介
"""
#import cv2
import numpy as np
import glob
import os
import openpyxl
import pprint
import copy

#exec('marker_2_x[i] = ari%dx[0][i]' % marker_2[i])

#excelファイルから配列を取り込む
wb = openpyxl.load_workbook('20190215_京大VS同志社_P3.xlsx')
wc = openpyxl.load_workbook('data0624_20190215_京大VS同志社_P3.xlsx')
sheet = wb['Sheet1']
c_sheet = wc['data']

#def get_value_list(t_2d):
#    return([[cell.value for cell in row] for row in t_2d])
#    
#    
#def get_list_2d(sheet, start_row, end_row, start_col, end_col):
#    return get_value_list(sheet.iter_cols(min_row=start_row,
#                                          max_row=end_row,
#                                          min_col=start_col,
#                                          max_col=end_col))
ari1x = []
ari1y = []
ari2x = []
ari2y = []
ari3x = []
ari3y = []
ari4x = []
ari4y = []
dfc1x = []
dfc1y = []
dfc2x = []
dfc2y = []
dfc3x = []
dfc3y = []
dfc4x = []
dfc4y = []
ballx = []
bally = []
ball_pos = []
pressure = []
passcource = []
for i, cell_obj in enumerate(list(c_sheet.rows)[0]):
    if i > 0:
        ball_pos.append(cell_obj.value)
for i, cell_obj in enumerate(list(c_sheet.rows)[1]):
    if i > 0:
        pressure.append(cell_obj.value)
for i, cell_obj in enumerate(list(c_sheet.rows)[2]):
    if i > 0:
        passcource.append(cell_obj.value)


for i, cell_obj in enumerate(list(sheet.columns)[3]):
    if i > 6:
            ari1x.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[4]):
    if i > 6:
            ari1y.append(cell_obj.value)
            
for i, cell_obj in enumerate(list(sheet.columns)[5]):
    if i > 6:
            ari2x.append(cell_obj.value)

for i, cell_obj in enumerate(list(sheet.columns)[6]):
    if i > 6:
            ari2y.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[7]):
    if i > 6:
            ari3x.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[8]):
    if i > 6:
            ari3y.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[9]):
    if i > 6:
            ari4x.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[10]):
    if i > 6:
            ari4y.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[11]):
    if i > 6:
            dfc1x.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[12]):
    if i > 6:
            dfc1y.append(cell_obj.value)
            
for i, cell_obj in enumerate(list(sheet.columns)[13]):
    if i > 6:
            dfc2x.append(cell_obj.value)

for i, cell_obj in enumerate(list(sheet.columns)[14]):
    if i > 6:
            dfc2y.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[15]):
    if i > 6:
            dfc3x.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[16]):
    if i > 6:
            dfc3y.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[17]):
    if i > 6:
            dfc4x.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[18]):
    if i > 6:
            dfc4y.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[19]):
    if i > 6:
            ballx.append(cell_obj.value)
for i, cell_obj in enumerate(list(sheet.columns)[20]):
    if i > 6:
            bally.append(cell_obj.value)

            
#ball_pos_pre = copy.copy(ball_pos)    
#listのlistとなることに注意する    
#start_row = 8
#end_row = 201
#ari1x = get_list_2d(sheet, start_row, end_row, 4, 4)
#ari1y = get_list_2d(sheet, start_row, end_row, 5, 5)
#ari2x = get_list_2d(sheet, start_row, end_row, 6, 6)
#ari2y = get_list_2d(sheet, start_row, end_row, 7, 7)
#ari3x = get_list_2d(sheet, start_row, end_row, 8, 8)
#ari3y= get_list_2d(sheet, start_row, end_row, 9, 9)
#ari4x= get_list_2d(sheet, start_row, end_row, 10, 10)
#ari4y = get_list_2d(sheet, start_row, end_row,11, 11)
#dfc1x= get_list_2d(sheet, start_row, end_row, 12, 12)
#dfc1y = get_list_2d(sheet, start_row, end_row, 13, 13)
#dfc2x = get_list_2d(sheet, start_row, end_row, 14, 14)
#dfc2y = get_list_2d(sheet, start_row, end_row, 15, 15)
#dfc3x= get_list_2d(sheet, start_row, end_row, 16, 16)
#dfc3y= get_list_2d(sheet, start_row, end_row, 17, 17)
#dfc4x= get_list_2d(sheet, start_row, end_row, 18, 18)
#dfc4y= get_list_2d(sheet, start_row, end_row, 19, 19)
#
#ballx = get_list_2d(sheet, start_row, end_row, 20, 20)
#bally = get_list_2d(sheet, start_row,end_row, 21, 21)
#
#4選手のボールとの距離を求め、それが閾値未満ならボール保持者のラベルを付ける
#もし2選手以上にラベルがあれば、近いほうを保持者とする
#近くに誰もいないときは0のまま次の列に移動する
#最後まで行う
#ball_pos_17 = copy.copy(ball_pos)

#ball_pos = np.zeros(len(ballx))
#dist_ball = 1.5
##for list_x,list_y in zip(ballx, bally):
#for i,(x1,y1) in enumerate(zip(ballx,bally)):    
#    far_to_ball_p1 =(x1-dfc1x[i])**2 + (y1-dfc1y[i])**2  
#    if far_to_ball_p1 <= dist_ball:
#        ball_pos[i] = 1
#    
#    far_to_ball_p2 =(x1-dfc2x[i])**2 + (y1-dfc2y[i])**2 
#    if far_to_ball_p2 <= dist_ball and far_to_ball_p2 == min(far_to_ball_p1,far_to_ball_p2):
#        ball_pos[i] = 2
#
#    far_to_ball_p3 =(x1-dfc3x[i])**2 + (y1-dfc3y[i])**2
#    if far_to_ball_p3 <= dist_ball and far_to_ball_p3 == min(far_to_ball_p1,far_to_ball_p2,far_to_ball_p3):
#        ball_pos[i] = 3
#
#    far_to_ball_p4 =(x1-dfc4x[i])**2 + (y1-dfc4y[i])**2              
#    if far_to_ball_p4 <= dist_ball and far_to_ball_p4 == min(far_to_ball_p1,far_to_ball_p2,far_to_ball_p3,far_to_ball_p4):
#        ball_pos[i] = 4
#
##ドリブル中にラベルが無くなるのを解決したい
##スルーの時にラベルがつくのを解決したい
##no_label = 0
##for i, dat in enumerate(ball_pos):
###    if dat != 0 and dat != ball_pos[i-1] and dat != ball_pos[i+1]:
###        ball_pos[i] = ball_pos[i+1]
##    if dat == 0 and no_label == 0:
##        pre_label = ball_pos[i-1]
##        no_label += 1
##    elif dat ==0 and no_label !=0:
##        no_label += 1
##    elif dat !=0 and no_label !=0:
##        pos_label = ball_pos[i]
##        if pre_label == pos_label:
##            for i0 in range(no_label):
##                if ball_pos[i-i0-1] == 0:
##                    ball_pos[i-i0-1] = pos_label
##        no_label = 0        
#
#        
#        
#        
##        step += 1
##    else:
##        if step > 0:
##            step = 0
##        tmp_list[i] += -1 + step
##        step += -1
##plus_counter = max(tmp_list)
##minus_counter = -min(tmp_list)
##print(plus_counter)
##print(minus_counter)
#        
#    
#    
